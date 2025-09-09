import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# -----------------------------
# 1. Setup paths
# -----------------------------
log_folder = r"C:\Users\sk4_l\OneDrive\Documents\Projects\Kani\logs"
output_excel = r"C:\Users\sk4_l\OneDrive\Documents\Projects\Kani\LEO_Satellites_Telemetry.xlsx"

print("🚀 Starting Satellite Telemetry Processing Pipeline...\n")

# -----------------------------
# 2. Parse logs into dataframes
# -----------------------------
def parse_log_line(line):
    timestamp_match = re.match(r"\[(.*?)\]", line)
    if not timestamp_match:
        return None
    timestamp = timestamp_match.group(1)
    data_str = line.split("]")[1].strip()
    kv_pairs = [kv.strip() for kv in data_str.split(",")]
    row = {"Timestamp": timestamp}
    for kv in kv_pairs:
        if ": " in kv:
            key, value = kv.split(": ", 1)
            row[key.strip()] = value.strip()
    return row

sheets = {}
files_processed = 0

print("📂 Step 1: Reading log files from folder...")
for filename in os.listdir(log_folder):
    if filename.endswith(".txt"):
        satellite_name = os.path.splitext(filename)[0]
        filepath = os.path.join(log_folder, filename)

        print(f"   🔍 Found log file: {satellite_name}.txt")
        with open(filepath, "r", encoding="utf-8") as f:
            lines = f.readlines()

        lines = [line.strip() for line in lines if line.strip()]
        parsed_data = [parse_log_line(line) for line in lines]
        parsed_data = [row for row in parsed_data if row is not None]

        if parsed_data:
            df = pd.DataFrame(parsed_data)
            df.columns = [c.replace(" ", "_")
                           .replace("-", "_")
                           .replace("°", "deg")
                           .replace("/", "_") for c in df.columns]
            sheets[satellite_name] = df
            print(f"   ✅ {len(df)} telemetry records cleaned and loaded for {satellite_name}")
            files_processed += 1
        else:
            print(f"   ⚠️ No valid data in {satellite_name} logs!")

print(f"\n📊 Total satellites processed: {files_processed}\n")

# -----------------------------
# 3. Write all data to Excel with pandas
# -----------------------------
print("📂 Step 2: Writing satellite data into Excel workbook...")
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    for satellite, df in sheets.items():
        df.to_excel(writer, sheet_name=satellite[:31], index=False)
        print(f"   📑 Data written for {satellite}")
print("   ✅ All satellite data written to Excel.\n")

# -----------------------------
# 4. Load and modify file with openpyxl
# -----------------------------
print("🎨 Step 3: Applying conditional highlights for Warnings and Criticals...")
wb = load_workbook(output_excel)

warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for sheet_name in wb.sheetnames:
    if sheet_name.lower() == "summary":
        continue
    print(f"   🟠 Checking {sheet_name} for anomalies...")
    ws = wb[sheet_name]
    status_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value and str(ws.cell(row=1, column=col).value).lower() == "status":
            status_col = col
            break
    if status_col is None:
        print(f"   ⚠️ Skipped {sheet_name}: No 'Status' column found.")
        continue

    count = {"Warning": 0, "Critical": 0}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value:
            val = str(cell.value).strip().lower()
            fill = None
            if val == "warning":
                fill = warning_fill
                count["Warning"] += 1
            elif val == "critical":
                fill = critical_fill
                count["Critical"] += 1
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
    print(f"   ✅ Highlighted {count['Warning']} warnings, {count['Critical']} criticals in {sheet_name}.")

# -----------------------------
# 5. Create Summary Sheet
# -----------------------------
print("\n📂 Step 4: Building summary sheet with satellite health data...")
summary_data = []
for sheet_name in wb.sheetnames:
    if sheet_name.lower() == "summary":
        continue
    ws = wb[sheet_name]
    status_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value and str(ws.cell(row=1, column=col).value).lower() == "status":
            status_col = col
            break
    if status_col is None:
        continue

    ok = warning = critical = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value:
            val = str(cell.value).strip().lower()
            if val == "ok":
                ok += 1
            elif val == "warning":
                warning += 1
            elif val == "critical":
                critical += 1

    summary_data.append({"Satellite": sheet_name, "OK": ok, "Warning": warning, "Critical": critical})
    print(f"   📊 {sheet_name} → OK={ok}, Warning={warning}, Critical={critical}")

if "Summary" in wb.sheetnames:
    del wb["Summary"]
summary_ws = wb.create_sheet("Summary")
summary_ws.append(["Satellite", "OK", "Warning", "Critical"])
for data in summary_data:
    summary_ws.append([data["Satellite"], data["OK"], data["Warning"], data["Critical"]])

wb.save(output_excel)
print("   ✅ Summary sheet created and saved.\n")

# -----------------------------
# 6. Create Stacked Bar Chart inside Excel
# -----------------------------
print("📈 Step 5: Generating stacked bar chart in Excel...")
summary_ws = wb["Summary"]

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Satellite Status Distribution"
chart.y_axis.title = 'Count'
chart.x_axis.title = 'Satellites'
chart.grouping = "stacked"
chart.overlap = 100

data = Reference(summary_ws, min_col=2, max_col=4, min_row=1, max_row=summary_ws.max_row)
categories = Reference(summary_ws, min_col=1, min_row=2, max_row=summary_ws.max_row)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

summary_ws.add_chart(chart, "F2")

wb.save(output_excel)
print(f"   ✅ Chart added successfully at {output_excel}")

print("\n🎉 Processing Complete: Satellite telemetry logs cleaned, summarized, highlighted, and visualized!")
